import pdb
import openpyxl # A Python library to read/write Excel 2010 xlsx/xlsm files
# workbook is the entire excel file (is the container for all parts of the 
# document such another  worksheets
'''
Workbook Access
Worksheet Access
'''
import pandas as pd
x = pd.read_excel('example.xlsx', header = None)

wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames # returns a list of worksheets
wb.worksheets # returns a list of worksheets
sheet = wb['Sheet1'] # to access a worksheets
sheet.max_row
sheet.max_column
sheet.title

# go and change the active sheet and press save button
# and try different sheets and verify active_sheet is stored.
active_sheet = wb.active # use the property for the active sheet

''' 
Cell Access 
*A*
1. using Column Letter + Row Number (A1 ...) via active sheet
2. use a property called 'value' to read and write
'''
print(sheet['A1'])
print(sheet['A1'].value)
print(sheet['B1'])
print(sheet['C1'].value)

'''
*B*
1. using row and column numbers .cell(row, column)
2. Don't use get_highest_row()/column(), but use the property (max_row, max_column) to find the laregest
'''
print(sheet.max_row, " ", sheet.max_column) # not get_highest_row(), not get_highest_column()
print(sheet.cell(1,1).value) # 1st index, not 0th index

#'Row 1, Column 2 is Apples'
for i in range(1, sheet.max_row + 1): # 1st index, not 0th index
    print(i, sheet.cell(row = i, column = 2).value)
    # print(i, sheet.cell(row = i, column = 2).value, sheet.cell(row = i, column = 3).value)    

for i in range(1, sheet.max_row + 1): # 1st index, not 0th index
    for j in range(1, sheet.max_column + 1): # 1st index, not 0th index    
        print(i, sheet.cell(row = i, column = j).value, end="")
    print("\n")
    # print(i, sheet.cell(row = i, column = 2).value, sheet.cell(row = i, column = 3).value)    


# -- Different ways to get an index (numeric) from the column letters or vice versa
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string 
# not from openpyxl.cell import get_column_letter, column_index_from_string

get_column_letter(234)
column_index_from_string('A')
column_index_from_string('DKD')
x, y = coordinate_from_string('A4') # returns ('A',4)
col = column_index_from_string(x) # returns 1
row = y

# -- You also can slice Worksheet objects to get all the Cell objects in a row, column, or 
#     rectangular area of the spreadsheet. 
#     Then you can loop over all the cells in the slice.
gen_obj = tuple(sheet['A1':'C3']) 
print(gen_obj[0][0].value)
print(gen_obj[0][1].value)
print(gen_obj[0][2].value)

# -- now they are 0th index since they are now stored in a tuple 
# pdb.set_trace()
#
        # A1 Hello World!
        # B1 Apples
        # C1 73
        # A2 2015-04-05 03:41:23
        # B2 Hello Jungle
        # C2 85
        # A3 2015-04-06 12:46:51
        # B3 Pears
        # C3 14
for range_of_objs in sheet['A1':'C3']: # set the range of the data block
    for cell_obj in range_of_objs: # reading one row at a time and store it into a tuple
        print(cell_obj.coordinate, cell_obj.value)

# You can use a generator (a type of iterator) using .iter_rows by providing the range
for value in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3,values_only=True):
    print(value)

# -- Finally, here is how to write the results to a file.
#  -- Here is how to change the title of each sheet
wb = openpyxl.load_workbook('example.xlsx')
wb.sheetnames
sheet = wb['Sheet1']
sheet = wb[wb.sheetnames[0]]
sheet = wb[wb.sheetnames[1]]
sheet = wb[wb.sheetnames[2]]

sheet.title
sheet.title = "Page1"

# Let's write some values
sheet['A1'] = 'Hello World!' # A1, B2 ==> []
sheet['A1'].value
sheet.cell(2, 2).value = 'Hello Jungle' # (1, 2) ==> ()
sheet.cell(2, 2).value
wb.save('example2.xlsx')

# Creating create_sheet()
wb.create_sheet()
wb.create_sheet(index = 0, title='Page 1')

# Removing sheets using wb.remove_sheet()
wb.remove(wb['Sheet'])
del wb['Page1']
sheet = wb['Sheet1'] 

