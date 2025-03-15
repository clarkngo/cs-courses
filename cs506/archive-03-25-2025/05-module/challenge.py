from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#create excel type item
wb = Workbook()
# select the active worksheet
ws = wb.active

counter = 0
# row index incremented by 1 for row
# col index incremented by 1 for col.
# increment the counter
# counter += 1
# ws[] = counter
for column in range(1,11):
    column_letter = get_column_letter(column)
    for row in range(1,11):
        counter = counter +1
        ws[column_letter + str(row)] = counter

# values incrementing by 1 by column (Challenge in class)
# for row in range(1,11):
#     for column in range(1,6):
#         column_letter = get_column_letter(column)        
#         counter = counter +1
#         ws[column_letter + str(row)] = counter

wb.save("challenge.xlsx")
