from bs4 import BeautifulSoup
import requests

with open('simple.html') as html_file:
    soup = BeautifulSoup(html_file, 'lxml') # lxml is a parser

# source = requests.get('simple.html').text # response object

print(soup) # this prints the entire object
print(dir(soup))
print(soup.prettify())

# -- print each HTML tag information
match = soup.title #title vs title.text tag
print(soup.title)
# -- print only the textual information within each tag inner most
print(soup.title.text)
print(soup.h1)
# -- print only the textual information within each tag inner most
print(soup.h1.text) 
print(soup.body)
print(soup.body.text)


# -- let's look at the div tag (child tags)
# The div tag is known as Division tag. The div tag is used in HTML to make 
# divisions of content in the web page like (text, images, header, footer, navigation bar, etc). 
# Div tag has both open(<div>) and closing (</div>) tag and it is mandatory 
# to close the tag. The Div is the most usable tag in web development 
# because it helps us to separate out data in the web page and we can create 
# a particular section for particular data or function in the web pages.
# The <div> tag defines a division or a section in an HTML document. 
# The <div> tag is used as a container for HTML elements - which is then styled with CSS or manipulated with JavaScript. 

print(soup.div) 
# -- this only prints the first instance, and so what we need to do is to actually find 
#    a particular tag string that matches using find() method.
# -- Let's find the footer which as a different class name.
target = soup.find('div', class_ = 'footer') # find the tag with matching class attributes
print(target)

article = soup.find('div', class_ = 'article')
print(article)
print()
print(article.h2)
print()
print(article.h2.a)
print()
print(article.h2.a.text)

# -- What if I want to print "summary"?
# -- Can you figure out?
summary = article.p.text
print(summary)

# -- What if you like to extract only the HTML files that are assigned under href?
soup.findAll('a')
# [<a href="article_1.html">Module 8: Web Scrapping</a>,
#  <a href="article_2.html">Module 9: Image Manipulation</a>,
#  <a href="article_1.html">Module 10: Data Visualization</a>]
for href in soup.findAll('a'):
    print(href.get('href')) # get the value assigned to href


# -- Ok, now what if I want to print all the class' information?
print(soup.findAll('div', class_ = 'article')) # find all articles in a "LIST" format
# [
# <div class="article">
# <h2><a href="article_1.html">Module 8: Web Scrapping</a></h2>
# <p>Web Scrapping is based on a thought that it is a type of DB, and certain information can be retrieve automatically.</p>
# </div>, 
# <div class="article">
# <h2><a href="article_2.html">Module 9: Image Manipulation</a></h2>
# <p>Image mainpulation deals with reading, writing, editing and displaying some of the standard image format. </p>
# </div>
# ] <------- LIST

# -- this means I can loop thru the list.
for a in soup.findAll('div', class_ = 'article'): # find all articles
    print(a)
    print("---------")

articles = soup.findAll('div', class_ = 'article') # find all articles
for article in articles:
    heading = article.h2.a.text
    print(heading)
    summary = article.p.text
    print(summary)
    print("-----------")


import pdb
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors

wb = Workbook()
ws = wb.active # select the active worksheet
#create excel type item

articles = soup.findAll('div', class_ = 'article') # find all articles
row = 1
col = 1
for article in articles:
    heading = article.h2.a.text 
    print(heading)
    ws.cell(row, col).font = Font(bold=True)
    ws.cell(row, col).font = Font(bold=True, color=colors.RED, size=20)    
    ws.cell(row, col).value = heading
    summary = article.p.text 
    print(summary)
    row = row + 1
    ws.cell(row, col).font = Font(size=16)    
    ws.cell(row, col).value = summary
    print("-----------")
    row = row + 1

wb.save(".\\final.xlsx")
